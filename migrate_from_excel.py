#!/usr/bin/env python3
"""Migrate data from Bidding_Tracker_Pro Excel to SQLite database."""
import sys, os
sys.path.insert(0, os.path.dirname(__file__))

from pathlib import Path
from openpyxl import load_workbook
from models import (
    get_db, init_db, safe_float, safe_date_str,
    DOCUMENTS, MILESTONES
)

BIDDING_DIR = Path(__file__).parent.parent.resolve()
TRACKER_NAMES = [
    "Bidding_Tracker_Pro_v4_updated.xlsx",
    "Bidding_Tracker_Pro_v4.xlsx",
    "Bidding_Tracker_Pro_v3.xlsx",
]
SKIP_SHEETS = {"Dashboard", "Doc Tracker", "Cost Analysis", "Company Docs"}


def find_tracker():
    for name in TRACKER_NAMES:
        p = BIDDING_DIR / name
        if p.exists():
            return p
    return None


def migrate():
    tracker = find_tracker()
    if not tracker:
        print(f"No tracker file found in {BIDDING_DIR}")
        return False

    print(f"Migrating from: {tracker.name}")
    init_db()
    conn = get_db()

    try:
        wb = load_workbook(tracker, data_only=True)
    except Exception as e:
        print(f"Cannot open tracker: {e}")
        return False

    project_sheets = [s for s in wb.sheetnames if s not in SKIP_SHEETS]
    migrated = 0

    for sn in project_sheets:
        ws = wb[sn]

        # Read project data (v4 cell references)
        name = str(ws['C5'].value or sn)
        deadline = safe_date_str(ws['C6'].value)
        win_score = safe_float(ws['C7'].value, 50)
        readiness = safe_float(ws['C8'].value, 10)
        agency = str(ws['C9'].value or 'TBD')
        status = str(ws['C10'].value or 'NEED MORE INFO')

        # Cost data
        labor = safe_float(ws['C20'].value)
        materials = safe_float(ws['C21'].value)
        equipment = safe_float(ws['C22'].value)
        subcontractors = safe_float(ws['C23'].value)
        overhead = safe_float(ws['C24'].value)
        profit_pct = safe_float(ws['C26'].value, 15)

        # Scope & contact
        scope = str(ws['B16'].value or '') if ws['B16'].value and str(ws['B16'].value) != "(Enter scope here)" else ''
        contact = str(ws['B30'].value or '') if ws['B30'].value and str(ws['B30'].value) != "(Enter contact info)" else ''
        notes = str(ws['B58'].value or '')

        # Determine folder name from sheet metadata
        folder_name = str(ws['B2'].value or '').replace('Folder: ', '') if ws['B2'].value else name

        # Insert project
        try:
            cursor = conn.execute("""
                INSERT OR REPLACE INTO projects
                (sheet_name, folder_name, name, deadline, win_score, readiness, agency, status,
                 scope, contact_info, notes, labor, materials, equipment, subcontractors, overhead, profit_pct)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (sn, folder_name, name, deadline, win_score, readiness, agency, status,
                  scope, contact, notes, labor, materials, equipment, subcontractors, overhead, profit_pct))
            project_id = cursor.lastrowid
        except Exception as e:
            print(f"  Error inserting {sn}: {e}")
            continue

        # Migrate documents (rows 34-45, col D=status, E=source)
        for di, doc_name in enumerate(DOCUMENTS):
            r = 34 + di
            doc_status = str(ws.cell(row=r, column=4).value or 'Pending')
            source = str(ws.cell(row=r, column=5).value or '')
            doc_notes = str(ws.cell(row=r, column=6).value or '')
            try:
                conn.execute("""
                    INSERT OR REPLACE INTO documents (project_id, doc_name, status, source_file, notes)
                    VALUES (?, ?, ?, ?, ?)
                """, (project_id, doc_name, doc_status, source, doc_notes))
            except:
                pass

        # Migrate milestones (rows 48-55, col D=date, E=status)
        for mi, mile_name in enumerate(MILESTONES):
            r = 48 + mi
            mile_date = safe_date_str(ws.cell(row=r, column=4).value)
            mile_status = str(ws.cell(row=r, column=5).value or 'Upcoming')
            try:
                conn.execute("""
                    INSERT OR REPLACE INTO milestones (project_id, milestone_name, milestone_date, status)
                    VALUES (?, ?, ?, ?)
                """, (project_id, mile_name, mile_date, mile_status))
            except:
                pass

        migrated += 1
        print(f"  Migrated: {sn} ({name})")

    conn.execute("INSERT INTO activity_log (action, detail) VALUES (?, ?)",
                 ("migration", f"Migrated {migrated} projects from {tracker.name}"))
    conn.commit()
    conn.close()
    wb.close()

    print(f"\nDone! Migrated {migrated} projects to SQLite.")
    return True


if __name__ == "__main__":
    migrate()
