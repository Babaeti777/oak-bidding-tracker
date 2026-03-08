#!/usr/bin/env python3
"""OAK BUILDERS LLC — Bidding Tracker Pro v4 Builder
Complete rebuild with enhanced dashboard, document scanning, folder links.
"""
import sys, os
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, DataBarRule, FormatObject
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter

BIDDING_DIR = Path("/sessions/laughing-elegant-pascal/mnt/Bidding")
OUTPUT = BIDDING_DIR / "Bidding_Tracker_Pro_v4.xlsx"

# ═══════════════════════════════════════════
# COLOR PALETTE & STYLES
# ═══════════════════════════════════════════
NAVY      = "0F2439"
DARK_BLUE = "1B3A5C"
MED_BLUE  = "2B5C8A"
STEEL     = "3D7AB5"
LIGHT_BLUE= "D6E8F7"
ACCENT_GOLD = "E8A838"
ACCENT_GREEN = "27AE60"
ACCENT_RED   = "E74C3C"
WHITE     = "FFFFFF"
NEAR_WHITE= "F8F9FA"
LIGHT_GRAY= "EDF0F2"
MED_GRAY  = "BDC3C7"
INPUT_BG  = "FFFDE7"
WARN_BG   = "FFF3CD"
OK_BG     = "D4EDDA"
ERR_BG    = "F8D7DA"

FILL_NAVY      = PatternFill("solid", fgColor=NAVY)
FILL_DARK      = PatternFill("solid", fgColor=DARK_BLUE)
FILL_MED       = PatternFill("solid", fgColor=MED_BLUE)
FILL_STEEL     = PatternFill("solid", fgColor=STEEL)
FILL_LIGHT_BLUE= PatternFill("solid", fgColor=LIGHT_BLUE)
FILL_WHITE     = PatternFill("solid", fgColor=WHITE)
FILL_NEAR_WHITE= PatternFill("solid", fgColor=NEAR_WHITE)
FILL_GRAY      = PatternFill("solid", fgColor=LIGHT_GRAY)
FILL_INPUT     = PatternFill("solid", fgColor=INPUT_BG)
FILL_WARN      = PatternFill("solid", fgColor=WARN_BG)
FILL_OK        = PatternFill("solid", fgColor=OK_BG)
FILL_ERR       = PatternFill("solid", fgColor=ERR_BG)
FILL_GOLD      = PatternFill("solid", fgColor=ACCENT_GOLD)
FILL_GREEN_LITE= PatternFill("solid", fgColor="C6EFCE")
FILL_RED_LITE  = PatternFill("solid", fgColor="FFC7CE")
FILL_YEL_LITE  = PatternFill("solid", fgColor="FFEB9C")

FONT_TITLE   = Font(name="Arial", bold=True, color=WHITE, size=16)
FONT_SUBTITLE= Font(name="Arial", color="B0C4DE", size=10)
FONT_SECTION = Font(name="Arial", bold=True, color=WHITE, size=12)
FONT_HEADER  = Font(name="Arial", bold=True, color=WHITE, size=10)
FONT_KPI_NUM = Font(name="Arial", bold=True, color=NAVY, size=22)
FONT_KPI_LBL = Font(name="Arial", color="5D6D7E", size=9)
FONT_LABEL   = Font(name="Arial", bold=True, size=10, color="2C3E50")
FONT_DATA    = Font(name="Arial", size=10, color="2C3E50")
FONT_INPUT   = Font(name="Arial", size=10, color="0000FF")
FONT_LINK    = Font(name="Arial", size=10, color="008000")
FONT_FORMULA = Font(name="Arial", size=10, color="000000")
FONT_ALERT_R = Font(name="Arial", bold=True, color="C0392B", size=10)
FONT_ALERT_G = Font(name="Arial", bold=True, color="27AE60", size=10)
FONT_SMALL   = Font(name="Arial", size=9, color="7F8C8D")

THIN = Side(style="thin", color="BDC3C7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_BOTTOM = Border(bottom=Side(style="medium", color=MED_BLUE))
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right", vertical="center", wrap_text=True)
TOP_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)

STATUS_OPTIONS = "YES,NOPE,NEED MORE INFO,MAYBE,PREPARING,SUBMITTED,NOT BIDDING"
DOC_OPTIONS = "Pending,In Progress,Done,N/A"
MILE_OPTIONS = "Upcoming,Complete,Missed,N/A"

DOCUMENTS = [
    "Bid Bond / Bid Security", "Performance Bond", "Payment Bond",
    "Certificate of Insurance", "Bid Form / Proposal Form", "Non-Collusion Affidavit",
    "MBE/WBE Compliance", "Safety Plan / OSHA Logs", "Financial Statements",
    "References / Past Projects", "Subcontractor List", "Project Schedule",
]
MILESTONES = [
    "Pre-Bid Meeting", "Site Visit", "RFI Deadline", "Bid Submission",
    "Award Announcement", "Contract Signing", "Notice to Proceed", "Project Completion",
]

# Company documents from DOCUMENT_INVENTORY.md
COMPANY_DOCS = [
    ("Insurance", "Certificate of Insurance (COI)", "Have", "Insurance/COI.pdf"),
    ("Insurance", "Policy Declarations", "Have", "Insurance/Insurance_Policy_Declarations.pdf"),
    ("Insurance", "Workers' Compensation Certificate", "Need", ""),
    ("Insurance", "Automobile Liability Certificate", "Need", ""),
    ("Insurance", "Professional Liability / E&O", "Need", ""),
    ("Insurance", "Umbrella / Excess Liability", "Need", ""),
    ("Bonding", "Sample Bid Bond (Old Republic Surety)", "Have", "Bonding/BOND SMALLWOOD.pdf"),
    ("Bonding", "Bonding Capacity Letter", "Need", ""),
    ("Licenses", "Articles of Organization", "Have", "Licenses/Article of Organization - OAK.pdf"),
    ("Licenses", "Letter of Good Standing", "Have", "Licenses/OAK_Builders_LLC_Letter_of_Good_Standing-v3.pdf"),
    ("Licenses", "State Registration", "Have", "Licenses/State Registration Combined-CERT.pdf"),
    ("Licenses", "Business License (BE20741387)", "Have", "Licenses/BE20741387-21648392-CERT.pdf"),
    ("Personnel", "Resume — Shayan Abadian", "Have", "Personnel/Shayan Abadian 05-21-2024 2.pdf"),
    ("Personnel", "COR Level III Certification", "Have", "Personnel/Abadian COR Cert Level III.pdf"),
    ("Personnel", "FAC P-PM Senior Level", "Have", "Personnel/FAC P-PM Senior Level.pdf"),
    ("Personnel", "PMI Certification", "Have", "Personnel/PMI Certfication.pdf"),
    ("Personnel", "Work Experience Summary", "Have", "Personnel/Work Experience.pdf"),
    ("Financial", "Financial Statement (Balance Sheet / P&L)", "Need", ""),
    ("Financial", "Bank Reference Letters (×2)", "Need", ""),
    ("Financial", "Federal Tax Clearance", "Need", ""),
    ("Financial", "EMR (Experience Modification Rate)", "Need", ""),
    ("Safety", "OSHA 300 Logs (3 Years)", "Need", ""),
    ("Safety", "Safety Program Manual", "Need", ""),
    ("Gov't", "SAM.gov Registration", "Need", ""),
    ("Gov't", "DUNS Number Confirmation", "Need", ""),
    ("Experience", "Past Projects List (3-5 Years)", "Need", ""),
    ("Experience", "Client Reference Letters (3-5)", "Need", ""),
    ("Company", "Capability Statement", "Need", ""),
]

def to_date(val):
    if val is None or str(val).strip().upper() == "TBD":
        return "TBD"
    try:
        return datetime.strptime(str(val).strip()[:10], "%Y-%m-%d")
    except:
        return val

# ── PROJECT DATA ──
# Each: (sheet_name, display_name, folder_name, deadline, win%, readiness%, agency, status)
PROJECTS = [
    ("1343", "Fire Alarm Control Replacement (Rebid)", "1343-Fire Alarm Control Replacement at Arlington County Courthouse Building", "2026-03-09", 30, 30, "Arlington County", "PREPARING"),
    ("Rockville", "Concrete & Brick Infrastructure Repair", "Rockville -AMintenance of concrete and Brick Infrastructure", "2026-02-25", 45, 45, "City of Rockville", "PREPARING"),
    ("1341", "Central Library Plaza Repairs & Waterproofing", "1341-Central Library Plaza Repairs and Waterproofing", "2026-02-27", 50, 55, "Arlington County", "PREPARING"),
    ("VDOT", "VDOT General Contracting IFB 160604", "VDOT - PM", "2026-03-03", 40, 40, "VDOT", "PREPARING"),
    ("N1361", "Equipment Bureau Overhead Door Replacement", "ITB No. RFQ N1361 Equipment Bureau Overhead Door Replacement", "2026-03-20", 35, 20, "Arlington County", "PREPARING"),
    ("RFP113271", "Owner's Rep for School Construction", "Fairfax - Owner Rep", "2026-03-26", 55, 45, "Fairfax County", "PREPARING"),
    ("PUR1797", "Day Reporting Center Flooring Replacement", "PUR-1797 (Day Reporting Center Flooring Replacement)- Bid Attachments", "TBD", 50, 40, "Prince William County", "PREPARING"),
    ("GPO-FM", "GPO FM Modernization Construction", "Government Publishing Office FM Modernization Construction", "TBD", 35, 25, "US GPO", "PREPARING"),
    ("PUR1799", "Construction/Building Inspection Services", "PUR1779-CONSTRUCTIONBUILDING INSPECTION SERVICES", "TBD", 30, 15, "Prince William County", "NEED MORE INFO"),
    ("TaxCourt", "U.S. Tax Court Interior Paint & Signage", "U.S. Tax Court Interior Paint project", "TBD", 40, 15, "U.S. Tax Court", "NEED MORE INFO"),
    ("BPM055125", "Accordion Wall Divider — Springfield Hospital", "0000-Bid submitted/BPM055125 - Salomon Building Installation of Accordion Wall Divider for Springfield Hospital Center IFB", "2026-02-19", 60, 90, "MD DGS", "SUBMITTED"),
    ("GWRC", "GO Virginia Region 6 Project Admin", "0000-Bid submitted/GWRC - GO VIRGINIA", "2026-01-30", 45, 80, "GWRC", "SUBMITTED"),
    ("TuneUp", "TuneUp Salon Tenant Improvement", "0000-Bid submitted/Tune Up - 43800", "TBD", 70, 85, "Private Owner", "SUBMITTED"),
]
NUM_PROJ = len(PROJECTS)

def sc(ws, r, c, val, font=None, fill=None, align=None, border=None, nf=None):
    cell = ws.cell(row=r, column=c, value=val)
    if font: cell.font = font
    if fill: cell.fill = fill
    if align: cell.alignment = align
    if border: cell.border = border
    if nf: cell.number_format = nf
    return cell

def fill_range(ws, r1, c1, r2, c2, fill=None, border=None, font=None, align=None):
    for r in range(r1, r2+1):
        for c in range(c1, c2+1):
            cell = ws.cell(row=r, column=c)
            if fill: cell.fill = fill
            if border: cell.border = border
            if font: cell.font = font
            if align: cell.alignment = align

def make_folder_link(folder_name):
    return folder_name

# ═══════════════════════════════════════════
# DASHBOARD
# ═══════════════════════════════════════════
def build_dashboard(wb):
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_properties.tabColor = NAVY
    ws.sheet_view.showGridLines = False

    # Column widths
    widths = {"A": 3, "B": 4, "C": 32, "D": 14, "E": 10, "F": 10, "G": 10,
              "H": 14, "I": 18, "J": 14, "K": 10, "L": 10, "M": 36}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # ── HEADER (Rows 1-3) ──
    ws.merge_cells("A1:M1")
    sc(ws, 1, 1, "  OAK BUILDERS LLC — BIDDING COMMAND CENTER", FONT_TITLE, FILL_NAVY,
       Alignment(horizontal="left", vertical="center"))
    ws.row_dimensions[1].height = 50
    fill_range(ws, 1, 1, 1, 13, FILL_NAVY)

    ws.merge_cells("A2:M2")
    sc(ws, 2, 1, '=TEXT(TODAY(),"  Updated: mmmm d, yyyy")', FONT_SUBTITLE, FILL_NAVY,
       Alignment(horizontal="left", vertical="center"))
    ws.row_dimensions[2].height = 22
    fill_range(ws, 2, 1, 2, 13, FILL_NAVY)

    ws.row_dimensions[3].height = 8
    fill_range(ws, 3, 1, 3, 13, FILL_WHITE)

    # ── KPI SECTION (Rows 4-7) ──
    ws.merge_cells("B4:M4")
    sc(ws, 4, 2, "KEY PERFORMANCE INDICATORS", FONT_SECTION, FILL_MED, CENTER)
    fill_range(ws, 4, 2, 4, 13, FILL_MED)
    ws.row_dimensions[4].height = 28

    # KPI boxes: 6 metrics across columns B-M (2 cols each)
    kpi_data = [
        ("B5", "C5", f"={NUM_PROJ}", "TOTAL BIDS"),
        ("D5", "E5", f'=COUNTIFS(J{10}:J{9+NUM_PROJ},"<>NOT BIDDING",J{10}:J{9+NUM_PROJ},"<>NOPE")', "ACTIVE"),
        ("F5", "G5", f'=SUMPRODUCT((J{10}:J{9+NUM_PROJ}<>"NOT BIDDING")*(J{10}:J{9+NUM_PROJ}<>"NOPE")*H{10}:H{9+NUM_PROJ})', "PIPELINE $"),
        ("H5", "I5", f'=COUNTIF(J{10}:J{9+NUM_PROJ},"SUBMITTED")', "SUBMITTED"),
        ("J5", "K5", f'=IF(COUNTIFS(J{10}:J{9+NUM_PROJ},"<>NOT BIDDING",J{10}:J{9+NUM_PROJ},"<>NOPE")=0,0,ROUND(SUMPRODUCT((J{10}:J{9+NUM_PROJ}<>"NOT BIDDING")*(J{10}:J{9+NUM_PROJ}<>"NOPE")*K{10}:K{9+NUM_PROJ})/COUNTIFS(J{10}:J{9+NUM_PROJ},"<>NOT BIDDING",J{10}:J{9+NUM_PROJ},"<>NOPE"),0))', "AVG DOC %"),
        ("L5", "M5", f'=COUNTIFS(E{10}:E{9+NUM_PROJ},"<8",E{10}:E{9+NUM_PROJ},">-1",J{10}:J{9+NUM_PROJ},"<>NOT BIDDING",J{10}:J{9+NUM_PROJ},"<>NOPE")', "DUE < 7 DAYS"),
    ]
    ws.row_dimensions[5].height = 48
    ws.row_dimensions[6].height = 22

    for i, (c1, c2, formula, label) in enumerate(kpi_data):
        r5c1 = int(c1[1:]); col1_idx = ord(c1[0]) - 64; col2_idx = ord(c2[0]) - 64
        ws.merge_cells(f"{c1[0]}5:{c2[0]}5")
        ws.merge_cells(f"{c1[0]}6:{c2[0]}6")
        kpi_fill = FILL_LIGHT_BLUE
        sc(ws, 5, col1_idx, formula, FONT_KPI_NUM, kpi_fill, CENTER)
        ws.cell(row=5, column=col2_idx).fill = kpi_fill
        sc(ws, 6, col1_idx, label, FONT_KPI_LBL, kpi_fill, CENTER)
        ws.cell(row=6, column=col2_idx).fill = kpi_fill
        if "PIPELINE" in label:
            ws.cell(row=5, column=col1_idx).number_format = '$#,##0'
            ws.cell(row=5, column=col1_idx).font = Font(name="Arial", bold=True, color=NAVY, size=18)
        if "AVG DOC" in label:
            ws.cell(row=5, column=col1_idx).number_format = '0"%"'

    ws.row_dimensions[7].height = 6

    # ── PROJECT STATUS BOARD (Rows 8+) ──
    ws.merge_cells("B8:M8")
    sc(ws, 8, 2, "PROJECT STATUS BOARD", FONT_SECTION, FILL_DARK, CENTER)
    fill_range(ws, 8, 2, 8, 13, FILL_DARK)
    ws.row_dimensions[8].height = 28

    headers = ["#", "Project Name", "Deadline", "Days Left", "Win %", "Ready %",
               "Est. Price", "Agency", "Decision", "Doc %", "Alert", "Folder"]
    for ci, h in enumerate(headers, 2):
        sc(ws, 9, ci, h, FONT_HEADER, FILL_MED, CENTER, BORDER)
    ws.row_dimensions[9].height = 26

    # Freeze panes at row 10 so header stays visible
    ws.freeze_panes = "B10"

    # ── PROJECT ROWS ──
    for pi, (sn, dname, folder, deadline, win, ready, agency, status) in enumerate(PROJECTS):
        r = 10 + pi
        alt = FILL_WHITE if pi % 2 == 0 else FILL_NEAR_WHITE
        ws.row_dimensions[r].height = 24

        sc(ws, r, 2, pi+1, FONT_DATA, alt, CENTER, BORDER)
        # C5=Name, C6=Deadline, C7=Win%, C8=Ready%, C9=Agency, C10=Decision, C13=Doc%, C27=Price
        sc(ws, r, 3, f"='{sn}'!C5", FONT_LINK, alt, LEFT, BORDER)
        c = sc(ws, r, 4, f"='{sn}'!C6", FONT_FORMULA, alt, CENTER, BORDER)
        c.number_format = 'YYYY-MM-DD'
        sc(ws, r, 5, f'=IF(\'{sn}\'!C6="TBD","—",\'{sn}\'!C6-TODAY())', FONT_FORMULA, alt, CENTER, BORDER)
        c = sc(ws, r, 6, f"='{sn}'!C7", FONT_FORMULA, alt, CENTER, BORDER)
        c.number_format = '0"%"'
        c = sc(ws, r, 7, f"='{sn}'!C8", FONT_FORMULA, alt, CENTER, BORDER)
        c.number_format = '0"%"'
        c = sc(ws, r, 8, f"='{sn}'!C27", FONT_FORMULA, alt, RIGHT, BORDER)
        c.number_format = '$#,##0'
        sc(ws, r, 9, f"='{sn}'!C9", FONT_DATA, alt, LEFT, BORDER)
        sc(ws, r, 10, f"='{sn}'!C10", FONT_FORMULA, alt, CENTER, BORDER)
        c = sc(ws, r, 11, f"='{sn}'!C13", FONT_FORMULA, alt, CENTER, BORDER)
        c.number_format = '0"%"'
        sc(ws, r, 12, f'=IF(J{r}="NOT BIDDING","—",IF(J{r}="NOPE","—",IF(E{r}="—","—",IF(E{r}<0,"EXPIRED",IF(E{r}<7,"URGENT",IF(E{r}<14,"SOON","OK"))))))',
           FONT_FORMULA, alt, CENTER, BORDER)

    last_proj_row = 9 + NUM_PROJ

    # Actually let me fix the layout. K is Decision, L is Doc%, M is Alert
    # Rewrite: B=# C=Name D=Deadline E=DaysLeft F=Win% G=Ready% H=Price I=Agency J=EMPTY K=Decision L=Doc% M=Alert
    # Wait, that doesn't work with the headers defined. Let me keep the original mapping:
    # Col 2=# 3=Name 4=Deadline 5=DaysLeft 6=Win% 7=Ready% 8=Price 9=Agency 10=Decision 11=Doc% 12=Alert 13=Folder

    # Fix: rewrite project rows with correct column mapping
    # Delete the incorrect row data and redo
    # Actually let me just correct the formulas above - col 10 should be Decision, 11=Doc%, 12=Alert, 13=Folder

    for pi, (sn, dname, folder, deadline, win, ready, agency, status) in enumerate(PROJECTS):
        r = 10 + pi
        alt = FILL_WHITE if pi % 2 == 0 else FILL_NEAR_WHITE
        # Col 13 = Folder path (for update_tracker.py to convert to hyperlinks)
        sc(ws, r, 13, folder, FONT_SMALL, alt, LEFT, BORDER)

    # ── Conditional Formatting ──
    proj_range_j = f"J10:J{last_proj_row}"
    proj_range_l = f"L10:L{last_proj_row}"
    proj_range_e = f"E10:E{last_proj_row}"

    # Decision column colors
    ws.conditional_formatting.add(proj_range_j,
        CellIsRule(operator="equal", formula=['"SUBMITTED"'], fill=PatternFill("solid", fgColor="C6EFCE"), font=Font(color="006100", bold=True)))
    ws.conditional_formatting.add(proj_range_j,
        CellIsRule(operator="equal", formula=['"YES"'], fill=PatternFill("solid", fgColor="C6EFCE"), font=Font(color="006100", bold=True)))
    ws.conditional_formatting.add(proj_range_j,
        CellIsRule(operator="equal", formula=['"NOPE"'], fill=PatternFill("solid", fgColor="FFC7CE"), font=Font(color="9C0006")))
    ws.conditional_formatting.add(proj_range_j,
        CellIsRule(operator="equal", formula=['"NOT BIDDING"'], fill=PatternFill("solid", fgColor="FFC7CE"), font=Font(color="9C0006")))
    ws.conditional_formatting.add(proj_range_j,
        CellIsRule(operator="equal", formula=['"PREPARING"'], fill=PatternFill("solid", fgColor="B4D7F5"), font=Font(color="1B4F72")))
    ws.conditional_formatting.add(proj_range_j,
        CellIsRule(operator="equal", formula=['"NEED MORE INFO"'], fill=PatternFill("solid", fgColor="FFEB9C"), font=Font(color="9C6500")))
    ws.conditional_formatting.add(proj_range_j,
        CellIsRule(operator="equal", formula=['"MAYBE"'], fill=PatternFill("solid", fgColor="E8DAEF"), font=Font(color="6C3483")))

    # Alert column
    alert_range = f"L10:L{last_proj_row}"
    ws.conditional_formatting.add(alert_range,
        CellIsRule(operator="equal", formula=['"EXPIRED"'], fill=PatternFill("solid", fgColor="E74C3C"), font=Font(color=WHITE, bold=True)))
    ws.conditional_formatting.add(alert_range,
        CellIsRule(operator="equal", formula=['"URGENT"'], fill=PatternFill("solid", fgColor="E67E22"), font=Font(color=WHITE, bold=True)))
    ws.conditional_formatting.add(alert_range,
        CellIsRule(operator="equal", formula=['"SOON"'], fill=PatternFill("solid", fgColor="F39C12"), font=Font(color="7D6608", bold=True)))
    ws.conditional_formatting.add(alert_range,
        CellIsRule(operator="equal", formula=['"OK"'], fill=PatternFill("solid", fgColor="27AE60"), font=Font(color=WHITE, bold=True)))

    # Days Left: red gradient
    ws.conditional_formatting.add(proj_range_e,
        CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor="FADBD8"), font=Font(color="C0392B", bold=True)))
    ws.conditional_formatting.add(proj_range_e,
        CellIsRule(operator="between", formula=["0", "7"], fill=PatternFill("solid", fgColor="FDEBD0"), font=Font(color="E67E22", bold=True)))

    # Doc % data bars
    doc_range = f"K10:K{last_proj_row}"
    ws.conditional_formatting.add(doc_range, DataBarRule(
        start_type="num", start_value=0, end_type="num", end_value=100,
        color="27AE60"))

    # Ready % data bars
    ready_range = f"G10:G{last_proj_row}"
    ws.conditional_formatting.add(ready_range, DataBarRule(
        start_type="num", start_value=0, end_type="num", end_value=100,
        color="3498DB"))

    # ── CHARTS SECTION ──
    chart_start = last_proj_row + 3

    # Hidden data for pie chart
    statuses = ["PREPARING", "SUBMITTED", "YES", "NOPE", "NOT BIDDING", "NEED MORE INFO", "MAYBE"]
    ws.cell(row=chart_start, column=15).value = "Status"
    ws.cell(row=chart_start, column=16).value = "Count"
    for si, st in enumerate(statuses):
        ws.cell(row=chart_start+1+si, column=15).value = st
        ws.cell(row=chart_start+1+si, column=16).value = f'=COUNTIF(J10:J{last_proj_row},"{st}")'

    pie = PieChart()
    pie.title = "Bids by Decision"
    pie.style = 10
    pie_data = Reference(ws, min_col=16, min_row=chart_start, max_row=chart_start+len(statuses))
    pie_cats = Reference(ws, min_col=15, min_row=chart_start+1, max_row=chart_start+len(statuses))
    pie.add_data(pie_data, titles_from_data=True)
    pie.set_categories(pie_cats)
    pie.width = 16
    pie.height = 12
    ws.add_chart(pie, f"B{chart_start}")

    # Bar chart: Bid values
    bar = BarChart()
    bar.type = "col"
    bar.title = "Estimated Bid Prices"
    bar.y_axis.title = "Amount ($)"
    bar.style = 10
    bar_data = Reference(ws, min_col=8, min_row=9, max_row=last_proj_row)
    bar_cats = Reference(ws, min_col=3, min_row=10, max_row=last_proj_row)
    bar.add_data(bar_data, titles_from_data=True)
    bar.set_categories(bar_cats)
    bar.shape = 4
    bar.width = 22
    bar.height = 12
    ws.add_chart(bar, f"H{chart_start}")

    return ws


# ═══════════════════════════════════════════
# PROJECT SHEETS
# ═══════════════════════════════════════════
def build_project_sheet(wb, sn, dname, folder, deadline, win, ready, agency, status):
    ws = wb.create_sheet(sn)
    ws.sheet_properties.tabColor = "3498DB" if status == "PREPARING" else ("27AE60" if status == "SUBMITTED" else "F39C12")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 18

    # ── ROW 1: Title Header ──
    ws.merge_cells("A1:F1")
    sc(ws, 1, 1, f"  {sn} — {dname}", FONT_TITLE, FILL_NAVY, Alignment(horizontal="left", vertical="center"))
    fill_range(ws, 1, 1, 1, 6, FILL_NAVY)
    ws.row_dimensions[1].height = 42

    # ── ROW 2: Folder Links ──
    ws.merge_cells("B2:C2")
    sc(ws, 2, 2, f"Folder: {folder}", FONT_SMALL, FILL_NEAR_WHITE, LEFT)
    ws.merge_cells("D2:F2")
    sc(ws, 2, 4, "CLAUDE Folder: (set by Update Tracker)", FONT_SMALL, FILL_NEAR_WHITE, LEFT)
    ws.row_dimensions[2].height = 20
    fill_range(ws, 2, 1, 2, 6, FILL_NEAR_WHITE)

    ws.row_dimensions[3].height = 6

    # ── KEY PROJECT DATA (Rows 4-10) ──
    ws.merge_cells("B4:C4")
    sc(ws, 4, 2, "KEY PROJECT DATA", FONT_SECTION, FILL_MED, CENTER)
    ws.cell(row=4, column=3).fill = FILL_MED
    ws.row_dimensions[4].height = 26

    fields = [
        ("Project Name", dname, None, None),
        ("Bid Deadline", to_date(deadline), None, "YYYY-MM-DD"),
        ("Win Score (%)", win, None, '0"%"'),
        ("Readiness (%)", ready, None, '0"%"'),
        ("Agency / Owner", agency, None, None),
        ("BIDDING Decision", status, None, None),
    ]
    for i, (label, val, _, nf) in enumerate(fields):
        r = 5 + i
        sc(ws, r, 2, label, FONT_LABEL, FILL_GRAY, LEFT, BORDER)
        cell = sc(ws, r, 3, val, FONT_INPUT, FILL_INPUT, LEFT, BORDER)
        if nf: cell.number_format = nf

    # Data validation for Decision
    dv = DataValidation(type="list", formula1=f'"{STATUS_OPTIONS}"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(ws["C10"])

    # ── Calculated Fields (Rows 12-13) ──
    ws.row_dimensions[11].height = 6
    sc(ws, 12, 2, "Days Until Deadline", FONT_LABEL, FILL_GRAY, LEFT, BORDER)
    sc(ws, 12, 3, '=IF(C6="TBD","TBD",C6-TODAY())', FONT_LINK, FILL_WHITE, LEFT, BORDER)

    sc(ws, 13, 2, "Document Readiness %", FONT_LABEL, FILL_GRAY, LEFT, BORDER)
    sc(ws, 13, 3,
       '=IF(COUNTIF(D34:D45,"Done")+COUNTIF(D34:D45,"In Progress")+COUNTIF(D34:D45,"Pending")=0,0,'
       'ROUND((COUNTIF(D34:D45,"Done"))/(COUNTIF(D34:D45,"Done")+COUNTIF(D34:D45,"In Progress")+COUNTIF(D34:D45,"Pending"))*100,0))',
       FONT_LINK, FILL_WHITE, LEFT, BORDER)
    ws.cell(row=13, column=3).number_format = '0"%"'

    # ── SCOPE (Rows 15-17) ──
    ws.row_dimensions[14].height = 6
    ws.merge_cells("B15:F15")
    sc(ws, 15, 2, "PROJECT SCOPE & DETAILS", FONT_SECTION, FILL_MED, CENTER)
    fill_range(ws, 15, 2, 15, 6, FILL_MED)
    ws.row_dimensions[15].height = 26

    ws.merge_cells("B16:F17")
    sc(ws, 16, 2, "(Enter scope description here)", FONT_DATA, FILL_INPUT, TOP_LEFT, BORDER)
    fill_range(ws, 16, 2, 17, 6, FILL_INPUT, BORDER)
    ws.row_dimensions[16].height = 30
    ws.row_dimensions[17].height = 30

    # ── COST BREAKDOWN (Rows 19-28) ──
    ws.row_dimensions[18].height = 6
    ws.merge_cells("B19:E19")
    sc(ws, 19, 2, "COST ESTIMATE BREAKDOWN", FONT_SECTION, FILL_MED, CENTER)
    fill_range(ws, 19, 2, 19, 5, FILL_MED)
    ws.row_dimensions[19].height = 26

    costs = ["Labor", "Materials", "Equipment", "Subcontractors", "Overhead / Indirect"]
    for ci, item in enumerate(costs):
        r = 20 + ci
        sc(ws, r, 2, item, FONT_LABEL, FILL_GRAY, LEFT, BORDER)
        c = sc(ws, r, 3, 0, FONT_INPUT, FILL_INPUT, LEFT, BORDER)
        c.number_format = '$#,##0'

    sc(ws, 25, 2, "TOTAL ESTIMATE", Font(name="Arial", bold=True, size=10, color=NAVY), PatternFill("solid", fgColor="D6E8F7"), LEFT, BORDER)
    c = sc(ws, 25, 3, "=SUM(C20:C24)", Font(name="Arial", bold=True, size=10, color="000000"), PatternFill("solid", fgColor="D6E8F7"), LEFT, BORDER)
    c.number_format = '$#,##0'

    sc(ws, 26, 2, "Target Profit %", FONT_LABEL, FILL_GRAY, LEFT, BORDER)
    c = sc(ws, 26, 3, 15, FONT_INPUT, FILL_INPUT, LEFT, BORDER)
    c.number_format = '0"%"'

    sc(ws, 27, 2, "Bid Price (w/ Profit)", FONT_LABEL, FILL_GRAY, LEFT, BORDER)
    c = sc(ws, 27, 3, '=IF(C25=0,0,C25*(1+C26/100))', Font(name="Arial", bold=True, size=11, color=ACCENT_GREEN), FILL_WHITE, LEFT, BORDER)
    c.number_format = '$#,##0'

    # ── CONTACT INFO (Rows 29-31) ──
    ws.row_dimensions[28].height = 6
    ws.merge_cells("B29:E29")
    sc(ws, 29, 2, "CONTACT INFORMATION", FONT_SECTION, FILL_MED, CENTER)
    fill_range(ws, 29, 2, 29, 5, FILL_MED)
    ws.row_dimensions[29].height = 26

    ws.merge_cells("B30:E31")
    sc(ws, 30, 2, "(Enter contact info)", FONT_DATA, FILL_INPUT, TOP_LEFT, BORDER)
    fill_range(ws, 30, 2, 31, 5, FILL_INPUT, BORDER)

    # ── DOCUMENT CHECKLIST (Rows 33-45) — with Source File column ──
    ws.row_dimensions[32].height = 6
    doc_headers = ["#", "Document", "Status", "Source File", "Notes"]
    for ci, h in enumerate(doc_headers):
        sc(ws, 33, 2+ci, h, FONT_HEADER, FILL_DARK, CENTER, BORDER)
    ws.row_dimensions[33].height = 26

    dv_doc = DataValidation(type="list", formula1=f'"{DOC_OPTIONS}"', allow_blank=True)
    ws.add_data_validation(dv_doc)

    for di, doc in enumerate(DOCUMENTS):
        r = 34 + di
        alt = FILL_WHITE if di % 2 == 0 else FILL_NEAR_WHITE
        sc(ws, r, 2, di+1, FONT_DATA, alt, CENTER, BORDER)
        sc(ws, r, 3, doc, FONT_DATA, alt, LEFT, BORDER)
        sc(ws, r, 4, "Pending", FONT_INPUT, FILL_INPUT, CENTER, BORDER)
        dv_doc.add(ws.cell(row=r, column=4))
        sc(ws, r, 5, "", FONT_SMALL, alt, LEFT, BORDER)  # Source File (auto-populated)
        sc(ws, r, 6, "", FONT_DATA, FILL_INPUT, LEFT, BORDER)  # Notes

    # Conditional formatting for doc status
    doc_range = f"D34:D{33+len(DOCUMENTS)}"
    ws.conditional_formatting.add(doc_range, CellIsRule(operator="equal", formula=['"Done"'], fill=FILL_GREEN_LITE, font=Font(color="006100", bold=True)))
    ws.conditional_formatting.add(doc_range, CellIsRule(operator="equal", formula=['"In Progress"'], fill=FILL_YEL_LITE, font=Font(color="9C6500")))
    ws.conditional_formatting.add(doc_range, CellIsRule(operator="equal", formula=['"Pending"'], fill=FILL_RED_LITE, font=Font(color="9C0006")))

    # ── MILESTONES (Rows 47-55) ──
    ws.row_dimensions[46].height = 6
    mile_headers = ["#", "Milestone", "Date", "Status"]
    for ci, h in enumerate(mile_headers):
        sc(ws, 47, 2+ci, h, FONT_HEADER, FILL_DARK, CENTER, BORDER)
    ws.row_dimensions[47].height = 26

    dv_mile = DataValidation(type="list", formula1=f'"{MILE_OPTIONS}"', allow_blank=True)
    ws.add_data_validation(dv_mile)

    for mi, mile in enumerate(MILESTONES):
        r = 48 + mi
        alt = FILL_WHITE if mi % 2 == 0 else FILL_NEAR_WHITE
        sc(ws, r, 2, mi+1, FONT_DATA, alt, CENTER, BORDER)
        sc(ws, r, 3, mile, FONT_DATA, alt, LEFT, BORDER)
        sc(ws, r, 4, "", FONT_INPUT, FILL_INPUT, CENTER, BORDER)
        sc(ws, r, 5, "Upcoming", FONT_INPUT, FILL_INPUT, CENTER, BORDER)
        dv_mile.add(ws.cell(row=r, column=5))

    ms_range = f"E48:E{47+len(MILESTONES)}"
    ws.conditional_formatting.add(ms_range, CellIsRule(operator="equal", formula=['"Complete"'], fill=FILL_GREEN_LITE, font=Font(color="006100")))
    ws.conditional_formatting.add(ms_range, CellIsRule(operator="equal", formula=['"Missed"'], fill=FILL_RED_LITE, font=Font(color="9C0006")))

    # ── NOTES (Rows 57-62) ──
    ws.row_dimensions[56].height = 6
    ws.merge_cells("B57:F57")
    sc(ws, 57, 2, "NOTES & ACTION ITEMS", FONT_SECTION, FILL_MED, CENTER)
    fill_range(ws, 57, 2, 57, 6, FILL_MED)
    ws.row_dimensions[57].height = 26

    ws.merge_cells("B58:F62")
    sc(ws, 58, 2, "", FONT_DATA, FILL_INPUT, TOP_LEFT, BORDER)
    fill_range(ws, 58, 2, 62, 6, FILL_INPUT, BORDER)

    return ws


# ═══════════════════════════════════════════
# DOC TRACKER SHEET
# ═══════════════════════════════════════════
def build_doc_tracker(wb):
    ws = wb.create_sheet("Doc Tracker")
    ws.sheet_properties.tabColor = "E67E22"
    ws.sheet_view.showGridLines = False

    ws.merge_cells(f"A1:{get_column_letter(3+NUM_PROJ)}1")
    sc(ws, 1, 1, "  DOCUMENT TRACKER — All Projects", FONT_TITLE, FILL_NAVY, Alignment(horizontal="left", vertical="center"))
    fill_range(ws, 1, 1, 1, 3+NUM_PROJ, FILL_NAVY)
    ws.row_dimensions[1].height = 42

    ws.row_dimensions[2].height = 6
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 4
    ws.column_dimensions[get_column_letter(3+NUM_PROJ)].width = 12

    # Headers
    sc(ws, 3, 2, "#", FONT_HEADER, FILL_DARK, CENTER, BORDER)
    sc(ws, 3, 3, "Document", FONT_HEADER, FILL_DARK, CENTER, BORDER)
    ws.column_dimensions["C"].width = 30

    for pi, (sn, dname, *_) in enumerate(PROJECTS):
        col = 4 + pi
        ws.column_dimensions[get_column_letter(col)].width = 14
        sc(ws, 3, col, sn, FONT_HEADER, FILL_MED, CENTER, BORDER)

    last_col = 4 + NUM_PROJ
    sc(ws, 3, last_col, "% Complete", FONT_HEADER, FILL_DARK, CENTER, BORDER)
    ws.column_dimensions[get_column_letter(last_col)].width = 12
    ws.row_dimensions[3].height = 28

    # Document rows
    for di, doc in enumerate(DOCUMENTS):
        r = 4 + di
        alt = FILL_WHITE if di % 2 == 0 else FILL_NEAR_WHITE
        sc(ws, r, 2, di+1, FONT_DATA, alt, CENTER, BORDER)
        sc(ws, r, 3, doc, FONT_DATA, alt, LEFT, BORDER)

        for pi, (sn, *_) in enumerate(PROJECTS):
            col = 4 + pi
            sc(ws, r, col, f"='{sn}'!D{34+di}", FONT_LINK, alt, CENTER, BORDER)

        # % Complete for this document across all projects
        first_col_l = get_column_letter(4)
        last_col_l = get_column_letter(3+NUM_PROJ)
        sc(ws, r, last_col,
           f'=IF(COUNTA({first_col_l}{r}:{last_col_l}{r})=0,0,ROUND(COUNTIF({first_col_l}{r}:{last_col_l}{r},"Done")/COUNTA({first_col_l}{r}:{last_col_l}{r})*100,0))',
           Font(name="Arial", bold=True, color="000000"), alt, CENTER, BORDER)
        ws.cell(row=r, column=last_col).number_format = '0"%"'

    # Conditional formatting for all status cells
    status_range = f"D4:{get_column_letter(3+NUM_PROJ)}{3+len(DOCUMENTS)}"
    ws.conditional_formatting.add(status_range, CellIsRule(operator="equal", formula=['"Done"'], fill=FILL_GREEN_LITE, font=Font(color="006100", bold=True)))
    ws.conditional_formatting.add(status_range, CellIsRule(operator="equal", formula=['"In Progress"'], fill=FILL_YEL_LITE, font=Font(color="9C6500")))
    ws.conditional_formatting.add(status_range, CellIsRule(operator="equal", formula=['"Pending"'], fill=FILL_RED_LITE, font=Font(color="9C0006")))

    # Separator
    sep_row = 4 + len(DOCUMENTS)
    ws.row_dimensions[sep_row].height = 6

    # Project completion row
    comp_row = sep_row + 1
    sc(ws, comp_row, 2, "", FONT_DATA, FILL_DARK, CENTER, BORDER)
    sc(ws, comp_row, 3, "PROJECT DOC READINESS", Font(name="Arial", bold=True, color=WHITE, size=10), FILL_DARK, LEFT, BORDER)
    for pi, (sn, *_) in enumerate(PROJECTS):
        col = 4 + pi
        c = sc(ws, comp_row, col, f"='{sn}'!C13",
               Font(name="Arial", bold=True, color="008000"), FILL_LIGHT_BLUE, CENTER, BORDER)
        c.number_format = '0"%"'

    # Data bars on completion row
    comp_range = f"D{comp_row}:{get_column_letter(3+NUM_PROJ)}{comp_row}"
    ws.conditional_formatting.add(comp_range, DataBarRule(start_type="num", start_value=0, end_type="num", end_value=100, color="27AE60"))

    ws.freeze_panes = "D4"
    return ws


# ═══════════════════════════════════════════
# COST ANALYSIS SHEET
# ═══════════════════════════════════════════
def build_cost_analysis(wb):
    ws = wb.create_sheet("Cost Analysis")
    ws.sheet_properties.tabColor = "2ECC71"
    ws.sheet_view.showGridLines = False

    ws.merge_cells(f"A1:{get_column_letter(4+NUM_PROJ)}1")
    sc(ws, 1, 1, "  COST ANALYSIS & COMPARISON", FONT_TITLE, FILL_NAVY, Alignment(horizontal="left", vertical="center"))
    fill_range(ws, 1, 1, 1, 4+NUM_PROJ, FILL_NAVY)
    ws.row_dimensions[1].height = 42

    ws.row_dimensions[2].height = 6
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 4
    ws.column_dimensions["C"].width = 22

    # Headers
    sc(ws, 3, 2, "#", FONT_HEADER, FILL_DARK, CENTER, BORDER)
    sc(ws, 3, 3, "Cost Category", FONT_HEADER, FILL_DARK, CENTER, BORDER)
    for pi, (sn, *_) in enumerate(PROJECTS):
        col = 4 + pi
        ws.column_dimensions[get_column_letter(col)].width = 14
        sc(ws, 3, col, sn, FONT_HEADER, FILL_MED, CENTER, BORDER)

    avg_col = 4 + NUM_PROJ
    sc(ws, 3, avg_col, "AVERAGE", FONT_HEADER, FILL_DARK, CENTER, BORDER)
    ws.column_dimensions[get_column_letter(avg_col)].width = 14
    ws.row_dimensions[3].height = 28

    cost_items = ["Labor", "Materials", "Equipment", "Subcontractors", "Overhead / Indirect"]
    for ci, item in enumerate(cost_items):
        r = 4 + ci
        alt = FILL_WHITE if ci % 2 == 0 else FILL_NEAR_WHITE
        sc(ws, r, 2, ci+1, FONT_DATA, alt, CENTER, BORDER)
        sc(ws, r, 3, item, FONT_LABEL, alt, LEFT, BORDER)
        for pi, (sn, *_) in enumerate(PROJECTS):
            col = 4 + pi
            c = sc(ws, r, col, f"='{sn}'!C{20+ci}", FONT_LINK, alt, CENTER, BORDER)
            c.number_format = '$#,##0'
        fcl = get_column_letter(4)
        lcl = get_column_letter(3+NUM_PROJ)
        c = sc(ws, r, avg_col, f"=IFERROR(AVERAGE({fcl}{r}:{lcl}{r}),0)", Font(name="Arial", bold=True, color="E67E22"), alt, CENTER, BORDER)
        c.number_format = '$#,##0'

    # Separator
    ws.row_dimensions[9].height = 6

    # Summary rows
    summary = [
        (10, "TOTAL ESTIMATE", "C25", True),   # =SUM(C20:C24)
        (11, "Bid Price w/ Profit", "C27", True),   # =IF(C25=0,0,C25*(1+C26/100))
        (12, "Target Profit %", "C26", False),  # hardcoded 15
    ]
    for r, label, cell_ref, is_currency in summary:
        sc(ws, r, 2, "", FONT_DATA, FILL_LIGHT_BLUE, CENTER, BORDER)
        sc(ws, r, 3, label, Font(name="Arial", bold=True, size=10, color=NAVY), FILL_LIGHT_BLUE, LEFT, BORDER)
        for pi, (sn, *_) in enumerate(PROJECTS):
            col = 4 + pi
            c = sc(ws, r, col, f"='{sn}'!{cell_ref}", Font(name="Arial", bold=True, color="008000"), FILL_LIGHT_BLUE, CENTER, BORDER)
            c.number_format = '$#,##0' if is_currency else '0"%"'
        fcl = get_column_letter(4)
        lcl = get_column_letter(3+NUM_PROJ)
        c = sc(ws, r, avg_col, f"=IFERROR(AVERAGE({fcl}{r}:{lcl}{r}),0)", Font(name="Arial", bold=True, color="E67E22"), FILL_LIGHT_BLUE, CENTER, BORDER)
        c.number_format = '$#,##0' if is_currency else '0"%"'

    # Profit amount row
    r = 13
    sc(ws, r, 2, "", FONT_DATA, FILL_OK, CENTER, BORDER)
    sc(ws, r, 3, "Profit Amount", Font(name="Arial", bold=True, size=10, color="006100"), FILL_OK, LEFT, BORDER)
    for pi in range(NUM_PROJ):
        col = 4 + pi
        c = sc(ws, r, col, f"={get_column_letter(col)}11-{get_column_letter(col)}10", Font(name="Arial", bold=True, color="006100"), FILL_OK, CENTER, BORDER)
        c.number_format = '$#,##0'
    c = sc(ws, r, avg_col, f"=IFERROR(AVERAGE({get_column_letter(4)}{r}:{get_column_letter(3+NUM_PROJ)}{r}),0)",
           Font(name="Arial", bold=True, color="E67E22"), FILL_OK, CENTER, BORDER)
    c.number_format = '$#,##0'

    ws.freeze_panes = "D4"
    return ws


# ═══════════════════════════════════════════
# COMPANY DOCUMENTS SHEET
# ═══════════════════════════════════════════
def build_company_docs(wb):
    ws = wb.create_sheet("Company Docs")
    ws.sheet_properties.tabColor = "8E44AD"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    sc(ws, 1, 1, "  COMPANY DOCUMENT INVENTORY", FONT_TITLE, FILL_NAVY, Alignment(horizontal="left", vertical="center"))
    fill_range(ws, 1, 1, 1, 7, FILL_NAVY)
    ws.row_dimensions[1].height = 42

    ws.merge_cells("A2:G2")
    sc(ws, 2, 1, "  Central repository: 1111-Claude/COMPANY_DOCUMENTS/", FONT_SUBTITLE, FILL_NAVY, Alignment(horizontal="left", vertical="center"))
    fill_range(ws, 2, 1, 2, 7, FILL_NAVY)
    ws.row_dimensions[2].height = 22

    ws.row_dimensions[3].height = 6
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 4
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 38
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 40
    ws.column_dimensions["G"].width = 22

    headers = ["#", "Category", "Document", "Status", "File Path", "Notes"]
    for ci, h in enumerate(headers):
        sc(ws, 4, 2+ci, h, FONT_HEADER, FILL_DARK, CENTER, BORDER)
    ws.row_dimensions[4].height = 26

    dv_status = DataValidation(type="list", formula1='"Have,Need,Expired,Ordered"', allow_blank=True)
    ws.add_data_validation(dv_status)

    current_cat = ""
    for di, (cat, doc_name, have_status, file_path) in enumerate(COMPANY_DOCS):
        r = 5 + di
        alt = FILL_WHITE if di % 2 == 0 else FILL_NEAR_WHITE

        # Category grouping
        show_cat = cat if cat != current_cat else ""
        if cat != current_cat:
            current_cat = cat

        sc(ws, r, 2, di+1, FONT_DATA, alt, CENTER, BORDER)
        sc(ws, r, 3, show_cat, Font(name="Arial", bold=True, size=10, color="2C3E50") if show_cat else FONT_DATA, alt, LEFT, BORDER)
        sc(ws, r, 4, doc_name, FONT_DATA, alt, LEFT, BORDER)
        sc(ws, r, 5, have_status, FONT_INPUT, FILL_INPUT, CENTER, BORDER)
        dv_status.add(ws.cell(row=r, column=5))
        sc(ws, r, 6, file_path, FONT_SMALL, alt, LEFT, BORDER)
        sc(ws, r, 7, "", FONT_DATA, FILL_INPUT, LEFT, BORDER)  # Notes

    # Conditional formatting
    last_doc_row = 4 + len(COMPANY_DOCS)
    status_range = f"E5:E{last_doc_row}"
    ws.conditional_formatting.add(status_range, CellIsRule(operator="equal", formula=['"Have"'], fill=FILL_GREEN_LITE, font=Font(color="006100", bold=True)))
    ws.conditional_formatting.add(status_range, CellIsRule(operator="equal", formula=['"Need"'], fill=FILL_RED_LITE, font=Font(color="9C0006", bold=True)))
    ws.conditional_formatting.add(status_range, CellIsRule(operator="equal", formula=['"Expired"'], fill=FILL_YEL_LITE, font=Font(color="9C6500", bold=True)))

    # Summary
    sum_row = last_doc_row + 2
    sc(ws, sum_row, 3, "TOTAL DOCUMENTS:", Font(name="Arial", bold=True, size=11, color=NAVY), FILL_LIGHT_BLUE, LEFT)
    sc(ws, sum_row, 4, len(COMPANY_DOCS), Font(name="Arial", bold=True, size=11, color=NAVY), FILL_LIGHT_BLUE, CENTER)

    sc(ws, sum_row+1, 3, "Documents Available:", FONT_LABEL, FILL_OK, LEFT)
    sc(ws, sum_row+1, 4, f'=COUNTIF(E5:E{last_doc_row},"Have")', Font(name="Arial", bold=True, color="006100"), FILL_OK, CENTER)

    sc(ws, sum_row+2, 3, "Documents Needed:", FONT_LABEL, FILL_ERR, LEFT)
    sc(ws, sum_row+2, 4, f'=COUNTIF(E5:E{last_doc_row},"Need")', Font(name="Arial", bold=True, color="9C0006"), FILL_ERR, CENTER)

    sc(ws, sum_row+3, 3, "Overall Readiness:", FONT_LABEL, FILL_LIGHT_BLUE, LEFT)
    sc(ws, sum_row+3, 4, f'=ROUND(COUNTIF(E5:E{last_doc_row},"Have")/COUNTA(E5:E{last_doc_row})*100,0)',
       Font(name="Arial", bold=True, size=14, color=NAVY), FILL_LIGHT_BLUE, CENTER)
    ws.cell(row=sum_row+3, column=4).number_format = '0"%"'

    ws.freeze_panes = "C5"
    return ws


# ═══════════════════════════════════════════
# MAIN BUILD
# ═══════════════════════════════════════════
def main():
    wb = Workbook()

    print("Building Dashboard...")
    build_dashboard(wb)

    print(f"Building {NUM_PROJ} project sheets...")
    for sn, dname, folder, deadline, win, ready, agency, status in PROJECTS:
        build_project_sheet(wb, sn, dname, folder, deadline, win, ready, agency, status)

    print("Building Doc Tracker...")
    build_doc_tracker(wb)

    print("Building Cost Analysis...")
    build_cost_analysis(wb)

    print("Building Company Docs inventory...")
    build_company_docs(wb)

    # Move Doc Tracker and Cost Analysis and Company Docs after project sheets
    # They should already be at the end, which is correct

    print(f"Saving to {OUTPUT}...")
    wb.save(OUTPUT)
    print(f"Saved! Sheets: {len(wb.sheetnames)}")

    # Count formulas
    formula_count = 0
    for sn in wb.sheetnames:
        ws = wb[sn]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and str(cell.value).startswith("="):
                    formula_count += 1
    print(f"Total formulas: {formula_count}")

if __name__ == "__main__":
    main()
